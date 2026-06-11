import io
import csv
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, status, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from database.connection import get_db
from database.models import Task, User, AuditLog, EvidenceFile, TaskReturn, TaskRejection
from backend.security import get_current_user, RoleChecker
from backend.utils import log_audit, format_duration

# ReportLab imports for PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

router = APIRouter(prefix="/api/reports", tags=["reports"])

def get_filtered_tasks_query(
    db: Session,
    category: Optional[str] = None,
    user_id: Optional[int] = None,
    role: Optional[str] = None,
    status: Optional[str] = None,
    completed_by: Optional[str] = None,  # "Manager", "NM Finance", "GM/CFO"
    returned: Optional[bool] = None,
    rejected: Optional[bool] = None,
    overdue: Optional[bool] = None,
    escalated: Optional[bool] = None,
    has_evidence: Optional[bool] = None,
    interval: Optional[str] = None,  # "Monthly", "Quarterly", "Half-Yearly", "Yearly"
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None
):
    query = db.query(Task)
    now = datetime.utcnow()

    if category:
        query = query.filter(Task.category == category)
    if user_id:
        query = query.filter(Task.created_by_id == user_id)
    if status:
        query = query.filter(Task.status == status)

    if role:
        if role == "Manager":
            query = query.filter(Task.status.in_(["Pending", "Returned to Initiator"]))
        elif role == "NM Finance":
            query = query.filter(Task.status == "Payroll Completed")
        elif role == "GM/CFO":
            query = query.filter(Task.status == "NM Finance Approved")

    if completed_by:
        if completed_by == "Manager":
            query = query.filter(Task.payroll_completed_by_id.isnot(None))
        elif completed_by == "NM Finance":
            query = query.filter(Task.nm_finance_approved_by_id.isnot(None))
        elif completed_by == "GM/CFO":
            query = query.filter(Task.gmcfo_approved_by_id.isnot(None))

    from sqlalchemy import select
    if returned is not None:
        return_subquery = select(TaskReturn.task_id)
        if returned:
            query = query.filter(Task.id.in_(return_subquery))
        else:
            query = query.filter(~Task.id.in_(return_subquery))

    if rejected is not None:
        rejection_subquery = select(TaskRejection.task_id)
        if rejected:
            query = query.filter(Task.id.in_(rejection_subquery))
        else:
            query = query.filter(~Task.id.in_(rejection_subquery))

    if overdue is not None:
        if overdue:
            query = query.filter(
                ((Task.actual_completion_date.isnot(None)) & (Task.actual_completion_date > Task.planned_due_date)) |
                ((Task.actual_completion_date.is_null()) & (Task.planned_due_date < now))
            )
        else:
            query = query.filter(
                ~(((Task.actual_completion_date.isnot(None)) & (Task.actual_completion_date > Task.planned_due_date)) |
                  ((Task.actual_completion_date.is_null()) & (Task.planned_due_date < now)))
            )

    if escalated is not None:
        seven_days_ago = now - timedelta(days=7)
        if escalated:
            query = query.filter(
                (Task.status != "GM/CFO Approved") & (Task.created_at <= seven_days_ago)
            )
        else:
            query = query.filter(
                ~((Task.status != "GM/CFO Approved") & (Task.created_at <= seven_days_ago))
            )

    if has_evidence is not None:
        if has_evidence:
            query = query.filter(Task.payroll_evidence_file_id.isnot(None))
        else:
            query = query.filter(Task.payroll_evidence_file_id.is_null())

    if interval:
        if interval == "Monthly":
            query = query.filter(Task.created_at >= now - timedelta(days=30))
        elif interval == "Quarterly":
            query = query.filter(Task.created_at >= now - timedelta(days=90))
        elif interval == "Half-Yearly":
            query = query.filter(Task.created_at >= now - timedelta(days=180))
        elif interval == "Yearly":
            query = query.filter(Task.created_at >= now - timedelta(days=365))

    if start_date:
        query = query.filter(Task.created_at >= start_date)
    if end_date:
        query = query.filter(Task.created_at <= end_date)

    return query

@router.get("/tasks")
def get_report_tasks(
    category: Optional[str] = None,
    user_id: Optional[int] = None,
    role: Optional[str] = None,
    status: Optional[str] = None,
    completed_by: Optional[str] = None,
    returned: Optional[bool] = None,
    rejected: Optional[bool] = None,
    overdue: Optional[bool] = None,
    escalated: Optional[bool] = None,
    has_evidence: Optional[bool] = None,
    interval: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    current_user: User = Depends(RoleChecker(["Administrator", "Auditor", "GM/CFO", "NM Finance"])),
    db: Session = Depends(get_db)
):
    query = get_filtered_tasks_query(
        db, category, user_id, role, status, completed_by,
        returned, rejected, overdue, escalated, has_evidence,
        interval, start_date, end_date
    )
    tasks = query.order_by(Task.created_at.desc()).all()
    now = datetime.utcnow()
    
    from backend.routes.tasks import map_task_details
    return [map_task_details(t, now) for t in tasks]

@router.get("/csv")
def export_csv(
    report_type: str = "tasks",  # "tasks", "audit", "evidence"
    category: Optional[str] = None,
    user_id: Optional[int] = None,
    role: Optional[str] = None,
    status: Optional[str] = None,
    completed_by: Optional[str] = None,
    returned: Optional[bool] = None,
    rejected: Optional[bool] = None,
    overdue: Optional[bool] = None,
    escalated: Optional[bool] = None,
    has_evidence: Optional[bool] = None,
    interval: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    current_user: User = Depends(RoleChecker(["Administrator", "Auditor", "GM/CFO", "NM Finance"])),
    db: Session = Depends(get_db)
):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    now = datetime.utcnow()
    
    if report_type == "tasks":
        query = get_filtered_tasks_query(
            db, category, user_id, role, status, completed_by,
            returned, rejected, overdue, escalated, has_evidence,
            interval, start_date, end_date
        )
        tasks = query.all()
        writer.writerow([
            "Task ID", "Title", "Description", "Department", "Category", "Status", 
            "Created By", "Created At", "Planned Due Date", "SLA Days", "SLA Status", "Rejection Count",
            "Payroll Completed By", "Payroll Completed At", 
            "Payroll Remarks", "Payroll Duration (hours)", "NM Finance Approved By", 
            "NM Finance Approved At", "NM Finance Remarks", "NM Finance Duration (hours)", 
            "GM/CFO Approved By", "GM/CFO Approved At", "GM/CFO Remarks", "GM/CFO Duration (hours)", 
            "Total Duration (hours)", "Archived"
        ])
        for t in tasks:
            planned = t.planned_due_date
            completed = t.actual_completion_date
            sla_status = "On Track"
            if planned:
                if completed:
                    if completed > planned:
                        sla_status = "Overdue"
                    else:
                        sla_status = "On Track"
                else:
                    if now > planned:
                        if (now - planned).days > 3 or (now - t.created_at).days > 7:
                            sla_status = "Critical"
                        else:
                            sla_status = "Overdue"
                    elif (planned - now).days <= 2:
                        sla_status = "Due Soon"

            writer.writerow([
                t.id, t.task_title, t.task_description or "", t.department, t.category, t.status,
                t.created_by.username if t.created_by else "System",
                t.created_at.strftime('%Y-%m-%d %H:%M:%S') if t.created_at else "",
                t.planned_due_date.strftime('%Y-%m-%d %H:%M:%S') if t.planned_due_date else "",
                t.sla_days or "",
                sla_status,
                t.rejection_count,
                t.payroll_completed_by.username if t.payroll_completed_by else "",
                t.payroll_completed_at.strftime('%Y-%m-%d %H:%M:%S') if t.payroll_completed_at else "",
                t.payroll_comments or "",
                round(t.payroll_processing_time / 3600.0, 2) if t.payroll_processing_time else "",
                t.nm_finance_approved_by.username if t.nm_finance_approved_by else "",
                t.nm_finance_approved_at.strftime('%Y-%m-%d %H:%M:%S') if t.nm_finance_approved_at else "",
                t.nm_finance_comments or "",
                round(t.nm_finance_processing_time / 3600.0, 2) if t.nm_finance_processing_time else "",
                t.gmcfo_approved_by.username if t.gmcfo_approved_by else "",
                t.gmcfo_approved_at.strftime('%Y-%m-%d %H:%M:%S') if t.gmcfo_approved_at else "",
                t.gmcfo_comments or "",
                round(t.gmcfo_processing_time / 3600.0, 2) if t.gmcfo_processing_time else "",
                round(t.total_completion_time / 3600.0, 2) if t.total_completion_time else "",
                t.is_archived
            ])
        filename = "tasks_report.csv"
        
    elif report_type == "audit":
        logs = db.query(AuditLog).order_by(AuditLog.timestamp.desc()).all()
        writer.writerow([
            "Log ID", "Username", "Timestamp", "Action", "Task ID", "IP Address", "Device", "Details", "Old Value", "New Value"
        ])
        for l in logs:
            writer.writerow([
                l.id, l.username, l.timestamp.strftime('%Y-%m-%d %H:%M:%S'), l.action_type, l.task_id or "",
                l.ip_address or "", l.device_info or "", l.details or "", l.old_value or "", l.new_value or ""
            ])
        filename = "audit_report.csv"
        
    elif report_type == "evidence":
        files = db.query(EvidenceFile).all()
        writer.writerow([
            "File ID", "Filename", "File Path", "SHA256 Hash", "Uploaded By", "Uploaded At", "Linked Task ID"
        ])
        for f in files:
            linked_task_id = f.task_reference[0].id if f.task_reference else ""
            writer.writerow([
                f.id, f.filename, f.filepath, f.file_hash,
                f.uploaded_by.username if f.uploaded_by else "Unknown",
                f.uploaded_at.strftime('%Y-%m-%d %H:%M:%S'),
                linked_task_id
            ])
        filename = "evidence_report.csv"
    else:
        raise HTTPException(status_code=400, detail="Invalid report type")

    log_audit(
        db=db,
        username=current_user.username,
        action_type="Report Downloads",
        user_id=current_user.id,
        details=f"Exported CSV report for '{report_type}'."
    )
    
    buffer.seek(0)
    return StreamingResponse(
        io.BytesIO(buffer.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/excel")
def export_excel(
    category: Optional[str] = None,
    user_id: Optional[int] = None,
    role: Optional[str] = None,
    status: Optional[str] = None,
    completed_by: Optional[str] = None,
    returned: Optional[bool] = None,
    rejected: Optional[bool] = None,
    overdue: Optional[bool] = None,
    escalated: Optional[bool] = None,
    has_evidence: Optional[bool] = None,
    interval: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    current_user: User = Depends(RoleChecker(["Administrator", "Auditor", "GM/CFO", "NM Finance"])),
    db: Session = Depends(get_db)
):
    query = get_filtered_tasks_query(
        db, category, user_id, role, status, completed_by,
        returned, rejected, overdue, escalated, has_evidence,
        interval, start_date, end_date
    )
    tasks = query.all()
    audit = db.query(AuditLog).all()
    users = db.query(User).all()
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: Tasks Registry
    ws_tasks = wb.active
    ws_tasks.title = "Tasks Registry"
    ws_tasks.append([
        "Task ID", "Title", "Department", "Category", "Status", "Created By", "Created At",
        "Planned Due Date", "SLA Status", "Rejection Count",
        "Payroll Completed By", "Payroll Processing Hours", "NM Finance Approved By",
        "NM Finance Processing Hours", "GM/CFO Approved By", "GM/CFO Processing Hours",
        "Total Processing Hours", "Archived"
    ])
    now = datetime.utcnow()
    for t in tasks:
        planned = t.planned_due_date
        completed = t.actual_completion_date
        sla_status = "On Track"
        if planned:
            if completed:
                if completed > planned:
                    sla_status = "Overdue"
                else:
                    sla_status = "On Track"
            else:
                if now > planned:
                    if (now - planned).days > 3 or (now - t.created_at).days > 7:
                        sla_status = "Critical"
                    else:
                        sla_status = "Overdue"
                elif (planned - now).days <= 2:
                    sla_status = "Due Soon"

        ws_tasks.append([
            t.id, t.task_title, t.department, t.category, t.status,
            t.created_by.username if t.created_by else "System",
            t.created_at.strftime('%Y-%m-%d %H:%M:%S') if t.created_at else "",
            t.planned_due_date.strftime('%Y-%m-%d %H:%M:%S') if t.planned_due_date else "",
            sla_status,
            t.rejection_count,
            t.payroll_completed_by.username if t.payroll_completed_by else "",
            round(t.payroll_processing_time / 3600.0, 2) if t.payroll_processing_time else 0.0,
            t.nm_finance_approved_by.username if t.nm_finance_approved_by else "",
            round(t.nm_finance_processing_time / 3600.0, 2) if t.nm_finance_processing_time else 0.0,
            t.gmcfo_approved_by.username if t.gmcfo_approved_by else "",
            round(t.gmcfo_processing_time / 3600.0, 2) if t.gmcfo_processing_time else 0.0,
            round(t.total_completion_time / 3600.0, 2) if t.total_completion_time else 0.0,
            t.is_archived
        ])
        
    # Sheet 2: User Productivity
    ws_users = wb.create_sheet(title="User Productivity")
    ws_users.append([
        "User ID", "Username", "Role", "Active", "Payroll Approvals Completed",
        "NM Finance Approvals Completed", "GM/CFO Approvals Completed", "Total Handled Steps"
    ])
    for u in users:
        s1_count = db.query(Task).filter(Task.payroll_completed_by_id == u.id).count()
        s2_count = db.query(Task).filter(Task.nm_finance_approved_by_id == u.id).count()
        s3_count = db.query(Task).filter(Task.gmcfo_approved_by_id == u.id).count()
        ws_users.append([
            u.id, u.username, u.role, u.is_active, s1_count, s2_count, s3_count, s1_count + s2_count + s3_count
        ])
        
    # Sheet 3: Audit Log Details
    ws_audit = wb.create_sheet(title="Audit Log Details")
    ws_audit.append([
        "Log ID", "Username", "Timestamp", "Action", "Task ID", "IP Address", "Device", "Details"
    ])
    for l in audit:
        ws_audit.append([
            l.id, l.username, l.timestamp.strftime('%Y-%m-%d %H:%M:%S'), l.action_type, l.task_id or "",
            l.ip_address or "", l.device_info or "", l.details or ""
        ])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    log_audit(
        db=db,
        username=current_user.username,
        action_type="Report Downloads",
        user_id=current_user.id,
        details="Exported Enterprise Excel Dashboard Workbook."
    )
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=enterprise_workflow_report.xlsx"}
    )

@router.get("/pdf")
def export_pdf(
    category: Optional[str] = None,
    user_id: Optional[int] = None,
    role: Optional[str] = None,
    status: Optional[str] = None,
    completed_by: Optional[str] = None,
    returned: Optional[bool] = None,
    rejected: Optional[bool] = None,
    overdue: Optional[bool] = None,
    escalated: Optional[bool] = None,
    has_evidence: Optional[bool] = None,
    interval: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    current_user: User = Depends(RoleChecker(["Administrator", "Auditor", "GM/CFO", "NM Finance"])),
    db: Session = Depends(get_db)
):
    # Generates a styled executive summary PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Corporate Styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#1E293B'),
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=30
    )
    
    section_title = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#4F46E5'),
        spaceBefore=15,
        spaceAfter=10
    )
    
    table_cell = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#334155'),
        leading=10
    )
    
    table_header = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white,
        leading=10
    )
    
    story = []
    
    # Header Section
    story.append(Paragraph("TaskTracker Pro - Enterprise Performance Report", title_style))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Target Auditor: {current_user.username} ({current_user.role})", subtitle_style))
    story.append(Spacer(1, 10))
    
    query = get_filtered_tasks_query(
        db, category, user_id, role, status, completed_by,
        returned, rejected, overdue, escalated, has_evidence,
        interval, start_date, end_date
    )
    tasks = query.all()

    # KPI metrics summary
    total_tasks = len(tasks)
    completed_tasks = len([t for t in tasks if t.status == "GM/CFO Approved"])
    pending_tasks = total_tasks - completed_tasks
    
    # SLA Compliance rate
    completed_tasks_list = [t for t in tasks if t.status == "GM/CFO Approved"]
    on_time = 0
    for t in completed_tasks_list:
        if t.planned_due_date and t.actual_completion_date:
            if t.actual_completion_date <= t.planned_due_date:
                on_time += 1
        else:
            on_time += 1
    sla_compliance_rate = (on_time / len(completed_tasks_list) * 100) if completed_tasks_list else 100.0
    
    # Rejections sum
    total_rejections = sum([t.rejection_count for t in tasks])
    
    kpi_data = [
        [
            Paragraph("<b>Total Tasks:</b>", table_cell), Paragraph(str(total_tasks), table_cell),
            Paragraph("<b>Completed Tasks:</b>", table_cell), Paragraph(str(completed_tasks), table_cell),
            Paragraph("<b>Pending Approval:</b>", table_cell), Paragraph(str(pending_tasks), table_cell)
        ],
        [
            Paragraph("<b>SLA Compliance:</b>", table_cell), Paragraph(f"{sla_compliance_rate:.1f}%", table_cell),
            Paragraph("<b>Total Rejections:</b>", table_cell), Paragraph(str(total_rejections), table_cell),
            Paragraph("", table_cell), Paragraph("", table_cell)
        ]
    ]
    kpi_table = Table(kpi_data, colWidths=[90, 80, 100, 80, 100, 80])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 20))
    
    # Table section: Active Pending Tasks
    story.append(Paragraph("Compliance Report Task Pipeline", section_title))
    
    table_data = [
        [
            Paragraph("ID", table_header),
            Paragraph("Title", table_header),
            Paragraph("Category", table_header),
            Paragraph("Workflow Status", table_header),
            Paragraph("Created At", table_header),
            Paragraph("Created By", table_header)
        ]
    ]
    
    for t in tasks[:25]:  # Limit to top 25 tasks for spacing constraints
        table_data.append([
            Paragraph(str(t.id), table_cell),
            Paragraph(t.task_title, table_cell),
            Paragraph(t.category, table_cell),
            Paragraph(t.status, table_cell),
            Paragraph(t.created_at.strftime('%Y-%m-%d'), table_cell),
            Paragraph(t.created_by.username if t.created_by else "System", table_cell)
        ])
        
    t_table = Table(table_data, colWidths=[30, 170, 90, 100, 70, 70])
    t_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    
    story.append(t_table)
    
    doc.build(story)
    buffer.seek(0)
    
    log_audit(
        db=db,
        username=current_user.username,
        action_type="Report Downloads",
        user_id=current_user.id,
        details="Generated Corporate PDF Performance Summary report."
    )
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=executive_compliance_summary.pdf"}
    )
