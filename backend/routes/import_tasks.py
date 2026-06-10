import io
from datetime import datetime
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request, status
from pydantic import BaseModel
from sqlalchemy.orm import Session
import openpyxl
from database.connection import get_db
from database.models import RecurringTaskMaster, User
from backend.security import RoleChecker, get_current_user
from backend.utils import log_audit

router = APIRouter(prefix="/api/import", tags=["import"])

class ImportRowError(BaseModel):
    row: int
    errors: List[str]

class ImportReport(BaseModel):
    success_count: int
    failure_count: int
    errors: List[ImportRowError]

@router.post("/recurring", response_model=ImportReport)
async def import_recurring_tasks_excel(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(RoleChecker(["Administrator", "GM/CFO"])),
    db: Session = Depends(get_db)
):
    # Verify file signature
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Invalid file format. Upload only Excel files (.xlsx or .xls)"
        )
        
    content = await file.read()
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Unable to read Excel workbook: {str(e)}"
        )
        
    if wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
    else:
        raise HTTPException(status_code=400, detail="Excel workbook has no sheets.")
        
    # Read headers
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Uploaded Excel sheet is empty.")
        
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    
    # Expected columns validation
    expected = ["Task Name", "Category", "Frequency", "Description", "Responsible Role", "Reminder Days", "Start Date", "End Date", "Priority"]
    for col in expected:
        if col not in headers:
            raise HTTPException(
                status_code=400,
                detail=f"Missing expected template column: '{col}'. Check headers format."
            )
            
    # Map header indices
    h_idx = {col: headers.index(col) for col in expected}
    
    success_count = 0
    failure_count = 0
    errors_list = []
    
    # Pre-fetch active users mapping role to first active user
    roles_list = ["Payroll Team", "NM Finance", "GM/CFO", "Administrator", "Auditor"]
    role_user_map = {}
    for role in roles_list:
        u = db.query(User).filter(User.role == role, User.is_active == True).first()
        if u:
            role_user_map[role] = u.id
            
    # Iterate records (row 2 onwards)
    for idx, row in enumerate(rows[1:], start=2):
        # Skip completely empty rows
        if not any(cell is not None for cell in row):
            continue
            
        row_errors = []
        
        # Read cell inputs
        task_name = row[h_idx["Task Name"]]
        category = row[h_idx["Category"]]
        frequency = row[h_idx["Frequency"]]
        description = row[h_idx["Description"]]
        resp_role = row[h_idx["Responsible Role"]]
        reminder_days = row[h_idx["Reminder Days"]]
        start_date_val = row[h_idx["Start Date"]]
        end_date_val = row[h_idx["End Date"]]
        priority = row[h_idx["Priority"]]
        
        # 1. Validate Task Name
        if not task_name or not str(task_name).strip():
            row_errors.append("Task Name is required.")
        else:
            task_name = str(task_name).strip()
            
        # 2. Validate Category
        valid_categories = ["Payroll", "Fund Accounting", "Factory Petty Cash", "Petty Cash", "Audit Schedules"]
        if not category or str(category).strip() not in valid_categories:
            row_errors.append(f"Category must be one of {valid_categories}")
        else:
            category = str(category).strip()
            
        # 3. Validate Frequency
        valid_freqs = ["Daily", "Weekly", "Monthly", "Quarterly", "Half-Yearly", "Yearly", "Every 2 Years"]
        if not frequency or str(frequency).strip() not in valid_freqs:
            row_errors.append(f"Frequency must be one of {valid_freqs}")
        else:
            frequency = str(frequency).strip()
            
        # 4. Validate Responsible Role & map user
        if not resp_role or str(resp_role).strip() not in roles_list:
            row_errors.append(f"Responsible Role must be one of {roles_list}")
        else:
            resp_role = str(resp_role).strip()
            if resp_role not in role_user_map:
                row_errors.append(f"No active user exists with the role '{resp_role}' to assign.")
                
        # 5. Validate Reminder Days
        parsed_reminder = 1
        if reminder_days is not None:
            try:
                parsed_reminder = int(reminder_days)
                if parsed_reminder < 0:
                    row_errors.append("Reminder Days cannot be negative.")
            except (ValueError, TypeError):
                row_errors.append("Reminder Days must be a valid integer.")
                
        # 6. Validate Start Date
        start_date = None
        if not start_date_val:
            row_errors.append("Start Date is required.")
        else:
            if isinstance(start_date_val, datetime):
                start_date = start_date_val
            else:
                try:
                    start_date = datetime.fromisoformat(str(start_date_val).replace("Z", ""))
                except ValueError:
                    try:
                        start_date = datetime.strptime(str(start_date_val).strip(), "%Y-%m-%d")
                    except ValueError:
                        row_errors.append("Start Date must be a valid date format (YYYY-MM-DD).")
                        
        # 7. Validate End Date (Optional)
        end_date = None
        if end_date_val:
            if isinstance(end_date_val, datetime):
                end_date = end_date_val
            else:
                try:
                    end_date = datetime.fromisoformat(str(end_date_val).replace("Z", ""))
                except ValueError:
                    try:
                        end_date = datetime.strptime(str(end_date_val).strip(), "%Y-%m-%d")
                    except ValueError:
                        row_errors.append("End Date must be a valid date format (YYYY-MM-DD) if provided.")
                        
        # 8. Validate Priority
        valid_priorities = ["High", "Normal", "Low"]
        if not priority or str(priority).strip() not in valid_priorities:
            row_errors.append(f"Priority must be one of {valid_priorities}")
        else:
            priority = str(priority).strip()
            
        # Collect or Save
        if row_errors:
            failure_count += 1
            errors_list.append(ImportRowError(row=idx, errors=row_errors))
        else:
            # Save configuration
            template = RecurringTaskMaster(
                task_name=task_name,
                category=category,
                description=str(description).strip() if description else "",
                responsible_person_id=role_user_map[resp_role],
                start_date=start_date,
                end_date=end_date,
                frequency=frequency,
                reminder_days=parsed_reminder,
                priority=priority,
                is_active=True,
                created_at=datetime.utcnow()
            )
            db.add(template)
            success_count += 1
            
    if success_count > 0:
        db.commit()
        log_audit(
            db=db,
            username=current_user.username,
            action_type="Task Creation", # Fits creating recurring configurations
            request=request,
            user_id=current_user.id,
            details=f"Bulk imported {success_count} recurring task templates from Excel."
        )
        
    return ImportReport(
        success_count=success_count,
        failure_count=failure_count,
        errors=errors_list
    )
